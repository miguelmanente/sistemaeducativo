# ==========================================================
# SGE - IMPORTADOR DE ASIGNACIONES DOCENTES
# Archivo: importador_asignaciones.py
#
# Lee archivos Excel desde la carpeta "importaciones",
# analiza los registros y prepara/importa asignaciones
# hacia la tabla "asignacion".
#
# Relaciones:
#
#   Excel DNI      -> profesores.id_docente
#   Excel Materia  -> materias.id_materia
#
# La tabla asignacion NO almacena DNI ni nombre de materia.
#
# Estados utilizados:
#
#   correcto   = asignación nueva, lista para importar
#   existente  = ya está en la base, no se importa
#   error      = dato inválido, impide la importación
#
# ==========================================================


# ==========================================================
# IMPORTACIONES
# ==========================================================

import os
from datetime import datetime

from openpyxl import load_workbook

from database import conectar


# ==========================================================
# CONFIGURACIÓN
# ==========================================================

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

CARPETA_IMPORTACIONES = os.path.join(
    BASE_DIR,
    "importaciones"
)


# ==========================================================
# COLUMNAS ESPERADAS DEL EXCEL
# ==========================================================

COLUMNAS_ESPERADAS = [
    "Marca temporal",
    "DNI",
    "Materia",
    "Cargo",
    "Curso",
    "Día",
    "Módulos",
    "Turno",
    "Hora de Entrada",
    "Hora de Salida",
    "Situación de Revista",
    "Toma de Posesión",
    "Cese",
    "Activo"
]


# ==========================================================
# CREAR CARPETA DE IMPORTACIONES
# ==========================================================

def crear_carpeta_importaciones():

    os.makedirs(
        CARPETA_IMPORTACIONES,
        exist_ok=True
    )


# ==========================================================
# OBTENER ARCHIVOS EXCEL
# ==========================================================

def obtener_archivos_excel():

    crear_carpeta_importaciones()

    archivos = []

    for nombre in os.listdir(
        CARPETA_IMPORTACIONES
    ):

        ruta = os.path.join(
            CARPETA_IMPORTACIONES,
            nombre
        )

        if not os.path.isfile(ruta):
            continue

        extension = os.path.splitext(
            nombre
        )[1].lower()

        if extension in (
            ".xlsx",
            ".xlsm"
        ):

            archivos.append(ruta)

    archivos.sort()

    return archivos


# ==========================================================
# LIMPIAR VALOR
# ==========================================================

def limpiar_valor(valor):

    if valor is None:
        return ""

    return str(valor).strip()


# ==========================================================
# NORMALIZAR TEXTO
# ==========================================================

def normalizar_texto(valor):

    return limpiar_valor(
        valor
    ).lower()


# ==========================================================
# NORMALIZAR DNI
# ==========================================================

def normalizar_dni(valor):

    if valor is None:
        return ""

    if isinstance(valor, float):

        if valor.is_integer():

            return str(
                int(valor)
            )

    texto = str(valor).strip()

    texto = texto.replace(
        ".",
        ""
    )

    texto = texto.replace(
        " ",
        ""
    )

    return texto


# ==========================================================
# CONVERTIR FECHA
# ==========================================================

def convertir_fecha(valor):

    if valor is None:
        return ""

    # ------------------------------------------------------
    # Fecha real de Excel
    # ------------------------------------------------------

    if isinstance(
        valor,
        datetime
    ):

        return valor.strftime(
            "%d/%m/%Y"
        )

    # ------------------------------------------------------
    # Texto
    # ------------------------------------------------------

    texto = str(valor).strip()

    if texto == "":
        return ""

    formatos = [
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%Y/%m/%d"
    ]

    for formato in formatos:

        try:

            fecha = datetime.strptime(
                texto,
                formato
            )

            return fecha.strftime(
                "%d/%m/%Y"
            )

        except ValueError:

            pass

    return texto


# ==========================================================
# CONVERTIR HORA
# ==========================================================

def convertir_hora(valor):

    if valor is None:
        return ""

    # ------------------------------------------------------
    # datetime
    # ------------------------------------------------------

    if isinstance(
        valor,
        datetime
    ):

        return valor.strftime(
            "%H:%M"
        )

    # ------------------------------------------------------
    # objetos que poseen strftime
    # ------------------------------------------------------

    if hasattr(
        valor,
        "strftime"
    ):

        try:

            return valor.strftime(
                "%H:%M"
            )

        except Exception:

            pass

    texto = str(
        valor
    ).strip()

    if texto == "":
        return ""

    formatos = [
        "%H:%M",
        "%H:%M:%S",
        "%I:%M %p",
        "%I:%M:%S %p"
    ]

    for formato in formatos:

        try:

            hora = datetime.strptime(
                texto,
                formato
            )

            return hora.strftime(
                "%H:%M"
            )

        except ValueError:

            pass

    return texto


# ==========================================================
# CONVERTIR MÓDULOS
# ==========================================================

def convertir_modulos(valor):

    if valor is None:
        return 0

    if isinstance(
        valor,
        int
    ):

        return valor

    if isinstance(
        valor,
        float
    ):

        if valor.is_integer():

            return int(
                valor
            )

    texto = str(
        valor
    ).strip()

    if texto == "":
        return 0

    try:

        return int(
            float(texto)
        )

    except (
        ValueError,
        TypeError
    ):

        return 0


# ==========================================================
# CONVERTIR ACTIVO
# ==========================================================

def convertir_activo(valor):

    if valor is None:
        return 1

    texto = str(
        valor
    ).strip().lower()

    if texto in (
        "0",
        "false",
        "no",
        "inactivo",
        "cesado"
    ):

        return 0

    return 1


# ==========================================================
# BUSCAR DOCENTE POR DNI
# ==========================================================

def buscar_docente(
    cursor,
    dni
):

    dni_normalizado = normalizar_dni(
        dni
    )

    if not dni_normalizado:
        return None

    cursor.execute(
        """
        SELECT id_docente
        FROM profesores
        WHERE REPLACE(
                  REPLACE(
                      TRIM(dni),
                      '.',
                      ''
                  ),
                  ' ',
                  ''
              ) = ?
        LIMIT 1
        """,
        (
            dni_normalizado,
        )
    )

    resultado = cursor.fetchone()

    if resultado:

        return resultado[0]

    return None


# ==========================================================
# BUSCAR MATERIA POR NOMBRE
# ==========================================================

def buscar_materia(
    cursor,
    nombre
):

    nombre_normalizado = normalizar_texto(
        nombre
    )

    if not nombre_normalizado:

        return None

    cursor.execute(
        """
        SELECT id_materia
        FROM materias
        WHERE LOWER(
                  TRIM(nombre)
              ) = ?
        LIMIT 1
        """,
        (
            nombre_normalizado,
        )
    )

    resultado = cursor.fetchone()

    if resultado:

        return resultado[0]

    return None


# ==========================================================
# BUSCAR ASIGNACIÓN EXISTENTE
# ==========================================================

def buscar_asignacion_existente(
    cursor,
    registro
):

    id_docente = registro[
        "id_docente"
    ]

    id_materia = registro[
        "id_materia"
    ]

    dia = registro[
        "dia"
    ]

    curso = registro[
        "curso"
    ]

    turno = registro[
        "turno"
    ]

    hentrada = registro[
        "hentrada"
    ]

    hsalida = registro[
        "hsalida"
    ]

    # ------------------------------------------------------
    # ASIGNACIÓN CON MATERIA
    # ------------------------------------------------------

    if id_materia is not None:

        cursor.execute(
            """
            SELECT id_asignacion
            FROM asignacion
            WHERE id_docente = ?
              AND id_materia = ?
              AND LOWER(TRIM(dia))
                    = LOWER(TRIM(?))
              AND LOWER(TRIM(curso))
                    = LOWER(TRIM(?))
              AND LOWER(TRIM(turno))
                    = LOWER(TRIM(?))
              AND TRIM(hentrada)
                    = TRIM(?)
              AND TRIM(hsalida)
                    = TRIM(?)
            LIMIT 1
            """,
            (
                id_docente,
                id_materia,
                dia,
                curso,
                turno,
                hentrada,
                hsalida
            )
        )

    # ------------------------------------------------------
    # CARGO SIN MATERIA
    # ------------------------------------------------------

    else:

        cursor.execute(
            """
            SELECT id_asignacion
            FROM asignacion
            WHERE id_docente = ?
              AND id_materia IS NULL
              AND LOWER(TRIM(dia))
                    = LOWER(TRIM(?))
              AND LOWER(TRIM(curso))
                    = LOWER(TRIM(?))
              AND LOWER(TRIM(turno))
                    = LOWER(TRIM(?))
              AND TRIM(hentrada)
                    = TRIM(?)
              AND TRIM(hsalida)
                    = TRIM(?)
            LIMIT 1
            """,
            (
                id_docente,
                dia,
                curso,
                turno,
                hentrada,
                hsalida
            )
        )

    resultado = cursor.fetchone()

    if resultado:

        return resultado[0]

    return None


# ==========================================================
# CLAVE DE ASIGNACIÓN
# ==========================================================

def clave_asignacion(
    registro
):

    return (
        registro[
            "id_docente"
        ],

        registro[
            "id_materia"
        ],

        normalizar_texto(
            registro["dia"]
        ),

        normalizar_texto(
            registro["curso"]
        ),

        normalizar_texto(
            registro["turno"]
        ),

        limpiar_valor(
            registro["hentrada"]
        ),

        limpiar_valor(
            registro["hsalida"]
        )
    )


# ==========================================================
# ANALIZAR ARCHIVO
# ==========================================================

def analizar_archivo(
    ruta_archivo
):

    resultado = {

        "archivo": ruta_archivo,

        "columnas_ok": False,

        "total": 0,

        "correctos": 0,

        "existentes": 0,

        "errores": 0,

        "advertencias": 0,

        "detalle_existentes": [],

        "detalle_errores": [],

        "detalle_advertencias": [],

        "registros": []
    }

    # ======================================================
    # VERIFICAR ARCHIVO
    # ======================================================

    if not os.path.isfile(
        ruta_archivo
    ):

        resultado[
            "detalle_errores"
        ].append(
            {
                "fila": 0,
                "errores": [
                    "El archivo no existe."
                ]
            }
        )

        resultado[
            "errores"
        ] = 1

        return resultado

    # ======================================================
    # ABRIR EXCEL
    # ======================================================

    try:

        libro = load_workbook(
            ruta_archivo,
            data_only=True
        )

    except Exception as e:

        resultado[
            "detalle_errores"
        ].append(
            {
                "fila": 0,
                "errores": [
                    "No se pudo abrir el archivo: "
                    f"{e}"
                ]
            }
        )

        resultado[
            "errores"
        ] = 1

        return resultado

    # ======================================================
    # PRIMERA HOJA
    # ======================================================

    hoja = libro.active

    # ======================================================
    # ENCABEZADOS
    # ======================================================

    encabezados = []

    for celda in hoja[1]:

        encabezados.append(
            limpiar_valor(
                celda.value
            )
        )

    # ======================================================
    # VERIFICAR COLUMNAS
    # ======================================================

    columnas_faltantes = []

    for columna in COLUMNAS_ESPERADAS:

        if columna not in encabezados:

            columnas_faltantes.append(
                columna
            )

    if columnas_faltantes:

        resultado[
            "detalle_errores"
        ].append(
            {
                "fila": 1,
                "errores": [
                    "Faltan columnas obligatorias: "
                    + ", ".join(
                        columnas_faltantes
                    )
                ]
            }
        )

        resultado[
            "errores"
        ] = 1

        return resultado

    resultado[
        "columnas_ok"
    ] = True

    # ======================================================
    # MAPA DE COLUMNAS
    # ======================================================

    indice_columnas = {}

    for indice, nombre in enumerate(
        encabezados
    ):

        indice_columnas[
            nombre
        ] = indice

    # ======================================================
    # CONEXIÓN
    # ======================================================

    conn = conectar()

    cursor = conn.cursor()

    # ======================================================
    # CONTROL DUPLICADOS DEL EXCEL
    # ======================================================

    claves_excel = set()

    # ======================================================
    # RECORRER FILAS
    # ======================================================

    for (
        numero_fila,
        fila
    ) in enumerate(
        hoja.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):

        # --------------------------------------------------
        # FILA VACÍA
        # --------------------------------------------------

        fila_vacia = True

        for valor in fila:

            if limpiar_valor(
                valor
            ):

                fila_vacia = False

                break

        if fila_vacia:

            continue

        resultado[
            "total"
        ] += 1

        errores_fila = []

        advertencias_fila = []

        # ==================================================
        # OBTENER DATOS
        # ==================================================

        dni = normalizar_dni(
            fila[
                indice_columnas["DNI"]
            ]
        )

        materia = limpiar_valor(
            fila[
                indice_columnas["Materia"]
            ]
        )

        cargo_excel = limpiar_valor(
            fila[
                indice_columnas["Cargo"]
            ]
        )

        curso = limpiar_valor(
            fila[
                indice_columnas["Curso"]
            ]
        )

        dia = limpiar_valor(
            fila[
                indice_columnas["Día"]
            ]
        )

        modulos = convertir_modulos(
            fila[
                indice_columnas["Módulos"]
            ]
        )

        turno = limpiar_valor(
            fila[
                indice_columnas["Turno"]
            ]
        )

        hentrada = convertir_hora(
            fila[
                indice_columnas[
                    "Hora de Entrada"
                ]
            ]
        )

        hsalida = convertir_hora(
            fila[
                indice_columnas[
                    "Hora de Salida"
                ]
            ]
        )

        situacion_revista = limpiar_valor(
            fila[
                indice_columnas[
                    "Situación de Revista"
                ]
            ]
        )

        toma_pos = convertir_fecha(
            fila[
                indice_columnas[
                    "Toma de Posesión"
                ]
            ]
        )

        fecha_cese = convertir_fecha(
            fila[
                indice_columnas[
                    "Cese"
                ]
            ]
        )

        activo = convertir_activo(
            fila[
                indice_columnas[
                    "Activo"
                ]
            ]
        )

        # ==================================================
        # VALIDAR DNI
        # ==================================================

        if not dni:

            errores_fila.append(
                "El DNI está vacío."
            )

        # ==================================================
        # BUSCAR DOCENTE
        # ==================================================

        id_docente = None

        if dni:

            id_docente = buscar_docente(
                cursor,
                dni
            )

            if id_docente is None:

                errores_fila.append(
                    f"El DNI '{dni}' no existe "
                    "en la tabla profesores."
                )

        # ==================================================
        # MATERIA / CARGO
        # ==================================================

        id_materia = None

        if materia:

            id_materia = buscar_materia(
                cursor,
                materia
            )

            if id_materia is None:

                errores_fila.append(
                    f"La materia '{materia}' "
                    "no existe en la tabla materias."
                )

            # ----------------------------------------------
            # Si existe materia:
            # cargo = Profesor
            # ----------------------------------------------

            cargo = "Profesor"

        else:

            id_materia = None

            cargo = cargo_excel

            if not cargo:

                errores_fila.append(
                    "La materia está vacía y "
                    "no se indicó ningún cargo."
                )

        # ==================================================
        # VALIDACIONES BÁSICAS
        # ==================================================

        if not dia:

            errores_fila.append(
                "El día está vacío."
            )

        if not turno:

            errores_fila.append(
                "El turno está vacío."
            )

        if not hentrada:

            errores_fila.append(
                "La hora de entrada está vacía."
            )

        if not hsalida:

            errores_fila.append(
                "La hora de salida está vacía."
            )

        if modulos < 0:

            errores_fila.append(
                "La cantidad de módulos "
                "no puede ser negativa."
            )

        # ==================================================
        # REGISTRO PREPARADO
        # ==================================================

        registro = {

            "fila_excel": numero_fila,

            "dni": dni,

            "id_docente": id_docente,

            "materia": materia,

            "id_materia": id_materia,

            "dia": dia,

            "cargo": cargo,

            "modulos": modulos,

            "curso": curso,

            "turno": turno,

            "hentrada": hentrada,

            "hsalida": hsalida,

            "situacion_revista":
                situacion_revista,

            "toma_pos": toma_pos,

            "fecha_cese": fecha_cese,

            "activo": activo,

            "estado": "correcto"
        }

        # ==================================================
        # SI NO HAY ERRORES:
        # CONTROLAR DUPLICADOS
        # ==================================================

        if not errores_fila:

            clave = clave_asignacion(
                registro
            )

            # ----------------------------------------------
            # DUPLICADO DENTRO DEL EXCEL
            # ----------------------------------------------

            if clave in claves_excel:

                registro[
                    "estado"
                ] = "existente"

                resultado[
                    "existentes"
                ] += 1

                resultado[
                    "detalle_existentes"
                ].append(
                    {
                        "fila":
                            numero_fila,

                        "mensaje":
                            "La asignación "
                            "está repetida "
                            "dentro del archivo Excel."
                    }
                )

            else:

                claves_excel.add(
                    clave
                )

                # ------------------------------------------
                # BUSCAR EN SQLITE
                # ------------------------------------------

                id_existente = (
                    buscar_asignacion_existente(
                        cursor,
                        registro
                    )
                )

                if id_existente is not None:

                    registro[
                        "estado"
                    ] = "existente"

                    resultado[
                        "existentes"
                    ] += 1

                    resultado[
                        "detalle_existentes"
                    ].append(
                        {
                            "fila":
                                numero_fila,

                            "id_asignacion":
                                id_existente,

                            "mensaje":
                                "La asignación "
                                "ya existe en "
                                "la base de datos."
                        }
                    )

        # ==================================================
        # SI HUBO ERROR REAL
        # ==================================================

        if errores_fila:

            registro[
                "estado"
            ] = "error"

            resultado[
                "errores"
            ] += 1

            resultado[
                "detalle_errores"
            ].append(
                {
                    "fila":
                        numero_fila,

                    "errores":
                        errores_fila
                }
            )

        # ==================================================
        # REGISTRO NUEVO
        # ==================================================

        elif registro[
            "estado"
        ] == "correcto":

            resultado[
                "correctos"
            ] += 1

        # ==================================================
        # GUARDAR REGISTRO
        # ==================================================

        resultado[
            "registros"
        ].append(
            registro
        )

    # ======================================================
    # CERRAR CONEXIÓN
    # ======================================================

    conn.close()

    return resultado


# ==========================================================
# IMPORTAR REGISTROS
# ==========================================================

def importar_registros(
    resultado
):

    # ======================================================
    # VALIDAR RESULTADO
    # ======================================================

    if resultado is None:

        return (
            False,
            "No se recibió ningún resultado de análisis."
        )

    if not resultado.get(
        "columnas_ok",
        False
    ):

        return (
            False,
            "El archivo no tiene las columnas esperadas."
        )

    # ======================================================
    # ERRORES REALES BLOQUEAN
    # ======================================================

    if resultado.get(
        "errores",
        0
    ) > 0:

        return (
            False,
            "No se puede importar porque "
            f"hay {resultado['errores']} "
            "registro(s) con errores."
        )

    # ======================================================
    # TOMAR SOLO REGISTROS NUEVOS
    # ======================================================

    registros_nuevos = []

    for registro in resultado.get(
        "registros",
        []
    ):

        if registro[
            "estado"
        ] == "correcto":

            registros_nuevos.append(
                registro
            )

    # ======================================================
    # SI NO HAY NUEVOS
    # ======================================================

    if not registros_nuevos:

        cantidad_existentes = resultado.get(
            "existentes",
            0
        )

        return (
            True,
            "No hay asignaciones nuevas "
            "para importar.\n\n"
            f"Asignaciones ya existentes: "
            f"{cantidad_existentes}"
        )

    # ======================================================
    # CONECTAR
    # ======================================================

    conn = conectar()

    cursor = conn.cursor()

    try:

        # ==================================================
        # SEGUNDA COMPROBACIÓN
        # JUSTO ANTES DE INSERTAR
        # ==================================================

        registros_a_insertar = []

        for registro in registros_nuevos:

            id_existente = (
                buscar_asignacion_existente(
                    cursor,
                    registro
                )
            )

            if id_existente is None:

                registros_a_insertar.append(
                    registro
                )

            else:

                # --------------------------------------------------
                # Puede haber cambiado la base entre el análisis
                # y el momento de importar.
                #
                # En ese caso simplemente no insertamos.
                # --------------------------------------------------

                registro[
                    "estado"
                ] = "existente"

        # ==================================================
        # INSERTAR
        # ==================================================

        for registro in registros_a_insertar:

            cursor.execute(
                """
                INSERT INTO asignacion (
                    id_docente,
                    id_materia,
                    dia,
                    cargo,
                    modulos,
                    curso,
                    turno,
                    hentrada,
                    hsalida,
                    situacion_revista,
                    toma_pos,
                    fecha_cese,
                    activo
                )
                VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?
                )
                """,
                (
                    registro[
                        "id_docente"
                    ],

                    registro[
                        "id_materia"
                    ],

                    registro[
                        "dia"
                    ],

                    registro[
                        "cargo"
                    ],

                    registro[
                        "modulos"
                    ],

                    registro[
                        "curso"
                    ],

                    registro[
                        "turno"
                    ],

                    registro[
                        "hentrada"
                    ],

                    registro[
                        "hsalida"
                    ],

                    registro[
                        "situacion_revista"
                    ],

                    registro[
                        "toma_pos"
                    ],

                    registro[
                        "fecha_cese"
                    ],

                    registro[
                        "activo"
                    ]
                )
            )

        # ==================================================
        # CONFIRMAR
        # ==================================================

        conn.commit()

        cantidad_importada = len(
            registros_a_insertar
        )

        cantidad_existentes = (
            resultado.get(
                "existentes",
                0
            )
            +
            (
                len(registros_nuevos)
                -
                cantidad_importada
            )
        )

        conn.close()

        return (
            True,
            "Importación finalizada correctamente.\n\n"
            f"Asignaciones nuevas importadas: "
            f"{cantidad_importada}\n"
            f"Asignaciones que ya existían: "
            f"{cantidad_existentes}\n"
            f"Errores: "
            f"{resultado.get('errores', 0)}"
        )

    except Exception as e:

        # ==================================================
        # DESHACER TODO
        # ==================================================

        conn.rollback()

        conn.close()

        return (
            False,
            "No se realizó la importación.\n\n"
            "La operación fue cancelada "
            "y la base de datos no fue modificada.\n\n"
            f"Detalle: {e}"
        )


# ==========================================================
# FIN DEL MÓDULO
# ==========================================================